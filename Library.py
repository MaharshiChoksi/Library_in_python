import datetime
import openpyxl.styles
from colorama import Fore, Style, init
import time
import random
import threading

global f_name, m_name, l_name, dob, age, gender, email, pn, acc_pin, AcId, u_city, address, T_C
init(autoreset=True)
ws = openpyxl.load_workbook("Database.xlsx")
wsa = ws.active
wk = openpyxl.load_workbook("Country-codes.xlsx")
wka = wk.active
log_status = "Logged out"
user_id = 0
Authorize = [9830183076]

ws_u = ws.worksheets[0]
ws_b = ws.worksheets[1]
ws_h = ws.worksheets[2]

"""
make a class where u can ask to login or register
then check if the user with user id logged in is manager or not.
if manager make instance of Lib_m class else lib_u
"""


class LibM:

    def __init__(self, user_id, log_status):
        self.user_id = user_id
        self.log_status = log_status
        self.a = 0
        self.bn = " "
        self.ba = " "
        self.mksr = " "
        self.brd = " "
        self.bpc = " "
        self.bnd = " "
        self.sr = 0000000
        self.status = " "
        self.bsr = 0000000
        self.cf = "x"
        self.bkr = " "
        self.bkar = " "
        self.bkpr = " "
        self.bkrq = " "
        self.bad = " "
        self.lend = " "
        self.lend_sr = 0000000
        self.us_charge = 0
        self.charge = 0
        self.welcome()

    def date_(self):
        try:
            while True:
                print(Style.DIM + Fore.CYAN + "Enter the release date in format (YYYY-MM-DD): ", end="")
                self.brd = input("")
                dt = datetime.date.today()
                if (int(self.brd[0:4]) <= int(dt.year)) and (len(self.brd) == 10):
                    if (self.brd[4] == "-") and (self.brd[7] == "-"):
                        if int(self.brd[5:7]) <= 12:
                            if (int(self.brd[8:10]) <= 30) or (int(self.brd[8:10]) <= 31):
                                return self.brd
                            else:
                                print(Style.DIM + Fore.RED + "Please enter valid Date...")
                        else:
                            print(Style.DIM + Fore.RED + "Please enter valid Month...")
                    else:
                        print(Style.DIM + Fore.RED + "Enter date with valid format")
                else:
                    if len(self.brd) != 10:
                        print(Style.DIM + Fore.RED + "Enter date with valid format")
                    else:
                        print(Style.DIM + Fore.RED + "Please enter valid year...")
        except Exception as e:
            print("Error!!!, ", e)
            self.date_()

    def new_book(self):
        while True:
            try:
                print(Style.DIM + Fore.CYAN + "Enter the name of Book: ", end="")
                self.bn = input("")
                print(Style.DIM + Fore.CYAN + "Enter the name of Author: ", end="")
                self.ba = input("")
                self.brd = self.date_()
                print(Style.DIM + Fore.CYAN + "Enter the publication House name: ", end="")
                self.bpc = input("")
                try:
                    while True:
                        self.sr = random.randint(1111111, 9999999)
                        i = 2
                        acc_list = []
                        while i < ws_b.max_row:
                            acc_no_database = ws_b.cell(i, 1)
                            acc_list.append(acc_no_database)
                            i += 1
                        if self.sr not in acc_list:
                            break
                except Exception as e:
                    print("Error!!!, ", e)
                self.status = "In Lib"
                return self.b_save()
            except Exception as e:
                print(Style.DIM + Fore.RED + f"Error: {e}")
                self.new_book()

    @staticmethod
    def l_history():
        try:
            print(Style.DIM + Fore.YELLOW + "\tLibrary Activity")
            for i in range(2, ws_h.max_row+1):
                print(Style.DIM + Fore.YELLOW + ws_h.cell(i, 1).value)
        except Exception as e:
            print(Style.DIM + Fore.RED + f"Error: {e}")

    def b_save(self):
        try:
            x = ws_b.max_row + 1
            ws_b.cell(x, 1).value = self.sr
            ws_b.cell(x, 2).value = self.bn
            ws_b.cell(x, 3).value = self.ba
            ws_b.cell(x, 4).value = self.brd
            ws_b.cell(x, 5).value = self.bpc
            ws_b.cell(x, 6).value = self.status
            y = ws_h.max_row + 1
            ws_h.cell(y, 1).value = f"New book '{self.bn}' of Author '{self.ba}' with sr.no '{self.sr}' has been added into Library"
            ws.save("Database.xlsx")
            print(Style.DIM + Fore.MAGENTA + "Your book is successfully added to library")
        except Exception as e:
            print(Style.DIM + Fore.RED + f"Error: {e}")

    def delete_book(self):
        try:
            print(Style.DIM + Fore.CYAN + "Name of books")
            for i in range(2, ws_b.max_row+1):
                print(Style.DIM + Fore.CYAN + f"\tBook-{i-1}: Sr.no {ws_b.cell(i, 1).value} Name: {ws_b.cell(i, 2).value}, Author: {ws_b.cell(i, 3).value}")
            print(Style.DIM + Fore.CYAN + "Enter name of book to delete: ", end="")
            self.bnd = input()
            print(Style.DIM + Fore.CYAN + "Enter name of author of the book: ", end="")
            self.bad = input()
            print(Style.DIM + Fore.CYAN + "Enter sr.no of the book: ", end="")
            self.bsr = input()
            for i in range(2, ws_b.max_row + 1):
                if self.bnd == str(ws_b.cell(i, 2).value) and self.bad == str(ws_b.cell(i, 3).value) and self.bsr == str(ws_b.cell(i, 1).value):
                    while True:
                        print(Style.DIM + Fore.RED + "Confirm, (Y) to delete this book: ", end="")
                        self.cf = input()
                        if self.cf.lower() == "y":
                            ws_b.delete_rows(i, 1)
                            y = ws_h.max_row + 1
                            ws_h.cell(y, 1).value = f"Book '{self.bnd}' of author '{self.bad}' with sr.no '{self.bsr}' has been deleted from Library"
                            return print(Style.DIM + Fore.RED + f"Book '{self.bnd}' is deleted from library'")
                        elif self.cf.lower() == "n":
                            return print(Style.DIM + Fore.RED + f"Book deleting of '{self.bnd}' was aborted!!!")
                        else:
                            print(Style.DIM + Fore.RED + "Enter valid input...")
            print(Style.DIM + Fore.RED + "Book name with this author not found")
        except Exception as e:
            print(Style.DIM + Fore.RED + f"Error: {e}")

    def make_changes(self):
        try:
            self.all_books()
            print(Style.DIM + Fore.CYAN + "Enter the sr.no of book that you want to make change: ", end="")
            self.mksr = input()
            x = ws_h.max_row+1
            for i in range(2, ws_b.max_row+1):
                if str(ws_b.cell(i, 1).value) == self.mksr:
                    while True:
                        print(Style.DIM + Fore.CYAN + "Enter the following option:\n1). change Book name\n2). Change Author name\n3). Change Publication name\n4). Change date\n0). End changing")
                        opt = int(input())
                        if opt == 1:
                            print(Style.DIM + Fore.CYAN + "Enter new Book name: ", end="")
                            new_bn = input()
                            ws_h.cell(x, 1).value = f"Book '{ws_b.cell(i,2).value}' has been renamed to '{new_bn}'"
                            ws_b.cell(i, 2).value = new_bn
                            break
                        elif opt == 2:
                            print(Style.DIM + Fore.CYAN + "Enter new author name: ", end="")
                            new_ba = input()
                            ws_h.cell(x, 1).value = f"Book author'{ws_b.cell(i,3).value}' has been renamed to '{new_ba}'"
                            ws_b.cell(i, 3).value = new_ba
                            break
                        elif opt == 3:
                            print(Style.DIM + Fore.CYAN + "Enter new publication name: ", end="")
                            new_pu = input()
                            ws_h.cell(x, 1).value = f"Book author'{ws_b.cell(i, 5).value}' has been renamed to '{new_pu}'"
                            ws_b.cell(i, 5).value = new_pu
                            break
                        elif opt == 4:
                            new_dt = self.date_()
                            ws_h.cell(x, 1).value = f"Book author'{ws_b.cell(i, 4).value}' has been renamed to '{new_dt}'"
                            ws_b.cell(i, 4).value = new_dt
                            break
                        elif opt == 0:
                            break
                        else:
                            print(Style.DIM + Fore.RED + "Choose appropriate value...")
                else:
                    if i == ws_b.max_row:
                        print(Style.DIM + Fore.RED + "Sr no not found")
        except Exception as e:
            print(Style.DIM + Fore.RED + f"Error: {e}")
            self.make_changes()

    def all_books(self):
        try:
            for i in range(1, ws_b.max_row + 1):
                for j in range(1, 7):
                    print(Style.BRIGHT + Fore.GREEN + str(ws_b.cell(i, j).value).ljust(20), end="\t")  # ljust will justify the text to left with appropriate format of 20 characters
                print("")
        except Exception as e:
            print(Style.DIM + Fore.RED + f"Error: {e}")
            self.all_books()

    def lending(self):
        try:
            print(Style.DIM + Fore.CYAN + "Enter book name to lend: ", end="")
            self.lend = input()
            print(Style.DIM + Fore.CYAN + "Enter book sr.no: ", end="")
            self.lend_sr = input()
            for i in range(2, ws_b.max_row + 1):
                if str(ws_b.cell(i, 2).value).lower() == self.lend and str(ws_b.cell(i, 1).value) == self.lend_sr:
                    if ws_b.cell(i, 6).value == "In Lib":
                        c = 2
                        while True:
                            if ws_u.cell(c, 1).value == user_id:
                                if int(ws_u.cell(c, 19).value) == 0 and str(ws_u.cell(c, 14).value) == "None":
                                    # add the book to the user page with due date, and it will show the charges amount per day after due date
                                    ws_u.cell(c, 13).value = f"{self.lend_sr},{self.lend}"
                                    ws_u.cell(c, 14).value = "Loaned"
                                    ws_u.cell(c, 15).value = str(datetime.date.today())
                                    ws_u.cell(c, 16).value = str(datetime.date.today() + datetime.timedelta(days=10))
                                    ws_u.cell(c, 17).value += f"You have lend {self.lend} book on {str(datetime.date.today())}"
                                    print(Style.DIM + Fore.RED + f"Lending Details\nBook: {self.lend_sr}, {self.lend}\nLending date: {str(datetime.date.today())}\nDue date: {str(datetime.date.today() + datetime.timedelta(days=10))}\nNOTE: Charges after due date is $1.00/day")
                                    # it will change the status of book to on loan
                                    ws_b.cell(i, 6).value = "On Loan"
                                    # also add it to the library activity
                                    ws_h.cell(ws_h.max_row+1, 1).value = f"User {user_id} has lend {self.lend_sr}, {self.lend} Book on {str(datetime.date.today())}."
                                    print(Style.DIM + Fore.RED + "Book Lending successful...")
                                    ws.save("Database.xlsx")
                                    break
                                else:
                                    if int(ws_u.cell(c, 19).value) != 0:
                                        print(Style.DIM + Fore.RED + f"You have {str(ws_u.cell(c, 19).value)}left on charges\nplease pay them first and then you can lend book...")
                                    elif str(ws_u.cell(c, 14).value) != "None":
                                        print(Style.DIM + Fore.RED + f"You have already lend a book..., please return it first....")
                                    break
                            else:
                                c += 1
                    else:
                        print(Style.DIM + Fore.RED + "Book is on loan by someone...\nplease check after few days...")
                else:
                    if i == ws_b.max_row+1:
                        print(Style.DIM + Fore.RED + "Book name with this sr no not found...")
                    else:
                        continue
        except Exception as e:
            print(Style.DIM + Fore.RED + f"Error: {e}")

    def returning(self):
        try:
            print(Style.DIM + Fore.CYAN + "Enter the name of book you want to return: ", end="")
            self.bkr = input()
            print(Style.DIM + Fore.CYAN + "Enter book sr.no: ", end="")
            self.lend_sr = input()
            # it will check if that book is lend or not, also the status of the user will aso be checked weather it is loaned or none
            x = 2
            while True:
                if ws_u.cell(x, 1).value == user_id:
                    if ws_u.cell(x, 14).value == "Loaned":
                        if str(ws_u.cell(x, 13).value).split(",")[0] == self.lend_sr and str(ws_u.cell(x, 13).value).split(",")[1] == self.bkr:
                            self.us_charge = str(self.give_charges())
                            if int(self.us_charge) == 0:
                                # book has been returned and set status of book to "In Lib", users status to "", add it to lib history
                                print(Style.DIM + Fore.RED + "Book has been returned to library without any charges...")
                                ws_u.cell(x, 19).value = "0"
                                ws_u.cell(x, 13).value = None
                                ws_u.cell(x, 14).value = None
                                ws_u.cell(x, 15).value = None
                                ws_u.cell(x, 16).value = None
                                ws_u.cell(x, 17).value += f"You have returned {self.lend_sr},{self.bkr} book on {str(datetime.date.today())} with no charges"
                                # changing status of book in lib
                                for i in range(1, ws_b.max_row+1):
                                    if str(ws_b.cell(i, 1).value) == str(self.lend_sr):
                                        ws_b.cell(i, 6).value = "In Lib"
                                ws_h.cell(ws_h.max_row+1, 1).value = f"user {user_id} returned {self.bkr}, {self.lend_sr} book on {datetime.date.today()}"
                                ws.save("Database.xlsx")
                            elif int(self.us_charge) != 0:
                                print(Style.DIM + Fore.RED + f"You have to pay ${self.us_charge} as charges and it has been added to your account")
                                ws_u.cell(x, 19).value = str(self.us_charge)
                                ws_u.cell(x, 13).value = None
                                ws_u.cell(x, 14).value = None
                                ws_u.cell(x, 15).value = None
                                ws_u.cell(x, 16).value = None
                                ws_u.cell(x, 17).value += f"You have returned {self.lend_sr},{self.bkr} book on {str(datetime.date.today())} with charges ${self.us_charge}"
                                # changing status of book in lib
                                for i in range(1, ws_b.max_row+1):
                                    if str(ws_b.cell(i, 1).value) == str(self.lend_sr):
                                        ws_b.cell(i, 6).value = "In Lib"
                                ws_h.cell(ws_h.max_row+1, 1).value = f"user {user_id} got ${self.us_charge} charges for returning {self.bkr} book late"
                                ws.save("Database.xlsx")
                            break
                        else:
                            print(Style.DIM + Fore.RED + "You haven't lend this book yet...")
                            break
                    else:
                        print(Style.DIM + Fore.RED + "You haven't loaned any book yet...")
                        break
                else:
                    x += 1
            # It will check if the due date have passed then it will add that amount of the charge to account.
            # while returning the book it will return the book and set the status of the book to in lib and users status to none and add it to users activity
            # it will add to library activity
        except Exception as e:
            print(Style.DIM + Fore.RED + f"Error: {e}")

    def give_charges(self):
        try:
            # this function will be activated when the user will return a book, and it will add the charges to users account and also in library activity
            x = 2
            while True:
                if ws_u.cell(x, 1).value == user_id:
                    if int(str(datetime.date.today()).split("-")[2]) > int(str(ws_u.cell(x , 16).value).split("-")[2]):
                        self.charge = int(str(datetime.date.today()).split("-")[2]) - int(str(ws_u.cell(x, 16).value).split("-")[2])
                        return self.charge
                    else:
                        self.charge = 0
                        return self.charge
                else:
                    x += 1
        except Exception as e:
            print(Style.DIM + Fore.RED + f"Error: {e}")

    def pay_charges(self):
        try:
            x = 2
            while True:
                if ws_u.cell(x, 1).value == user_id:
                    if int(ws_u.cell(x, 19).value) == 0:
                        print(Style.DIM + Fore.GREEN + "You don't have any charges...")
                        break
                    else:
                        print(Style.DIM + Fore.YELLOW + f"You have ${str(ws_u.cell(x, 19).value)} charges on your account...")
                        while True:
                            print("Please press (Y) to pay your charges: ", end="")
                            z = input()
                            if z.upper() == "Y":
                                print("You have successfully paid your charges")
                                ws_u.cell(x, 19).value = "0"
                                ws_h.cell(ws_h.max_row+1, 1).value = f"User {user_id} has paid ${str(ws_u.cell(x, 19).value)} charges"
                                ws.save("Database.xlsx")
                                break
                            elif z.upper() == "N":
                                print(Style.DIM + Fore.RED + "Charges payment aborted...")
                                break
                            else:
                                print(Style.DIM + Fore.RED + "Please enter valid choice...")
                        break
                else:
                    x += 1
        except Exception as e:
            print(Style.DIM + Fore.RED + f"Error: {e}")

    def lend_return(self):
        global ws
        try:
            while True:
                print(Style.DIM + Fore.CYAN + "Choose from the following option\n\t1). Lend book\n\t2). Return book\n\t3). Pay charges\n\t0). exit")
                choose = input()
                if int(choose) == 1:
                    self.lending()
                elif int(choose) == 2:
                    self.returning()
                elif int(choose) == 3:
                    self.pay_charges()
                elif int(choose) == 0:
                    break
                else:
                    print(Style.DIM + Fore.RED + "Choose appropriate option...")
                    ws.save("Database.xlsx")
                    ws = openpyxl.load_workbook("Database.xlsx")
        except Exception as e:
            print(Style.DIM + Fore.RED + f"Error: {e}")

    def request(self):
        try:
            print(Style.DIM + Fore.CYAN + "Enter the name of book you want: ", end="")
            self.bkr = input()
            print(Style.DIM + Fore.CYAN + "Enter the name of author of the book: ", end="")
            self.bkar = input()
            self.bkpr = self.date_()
            with open("request.txt", "a") as fx:
                fx.write(f"\nBook Request:\nBook name: {self.bkr}\nAuthor Name: {self.bkar}\nPublish year: {self.bkpr}\n")
            print(Style.DIM + Fore.RED + "Your request has been added...")
        except Exception as e:
            print(Style.DIM + Fore.RED + f"Error: {e}")
            self.request()

    def find_book_online(self):
        try:
            print("Enter the name of the book: ", end="")
            self.bkrq = input()
            if " " not in self.bkrq:
                print(f"Here is the link of the book that is available for the purchase: 'https://www.google.com/search?tbm=bks&q={self.bkrq}'")
            else:
                print(f"Here is the link of the book that is available for the purchase: 'https://www.google.com/search?tbm=bks&q={self.bkrq.replace(' ', '%20')}'")
        except Exception as e:
            print(Style.DIM + Fore.RED + f"Error: {e}")
            self.find_book_online()

    def welcome(self):
        global ws
        try:
            while True:
                print(Style.DIM + Fore.CYAN + "welcome to Uncensored library\n\t1) Add new book\n\t2) Delete book\n\t3) Make Changes\n\t4) See all books\n\t5) Lend or return book\n\t6) Request a book\n\t7) Find book online\n\t8) Activity in Library\n\t0) Log Out")
                self.a = int(input())
                if self.a not in [1, 2, 3, 4, 5, 6, 7, 8, 0]:
                    print(Style.DIM + Fore.RED + "Please choose valid option from above...")
                else:
                    if self.a == 1:
                        self.new_book()
                    elif self.a == 2:
                        self.delete_book()
                    elif self.a == 3:
                        self.make_changes()
                    elif self.a == 4:
                        self.all_books()
                    elif self.a == 5:
                        self.lend_return()
                    elif self.a == 6:
                        self.request()
                    elif self.a == 7:
                        self.find_book_online()
                    elif self.a == 8:
                        self.l_history()
                    elif self.a == 0:
                        Log.end()
                    ws.save("Database.xlsx")
                    ws = openpyxl.load_workbook("Database.xlsx")
        except Exception as e:
            print(Style.DIM + Fore.RED + f"Error: {e}")
            self.welcome()


class LibU(LibM):

    def acc_info(self):
        try:
            t = 2
            while True:
                if ws_u.cell(t, 1).value == user_id:
                    for y in range(1, ws_u.max_column+1):
                        if ws_u.cell(t, y).value != "None": print(f"{ws_u.cell(1, y).value}: {ws_u.cell(t, y).value}")
                    break
                else:
                    t += 1
        except Exception as e:
            print(Style.DIM + Fore.RED + f"Error: {e}")

    def welcome(self):
        try:
            global ws
            while True:
                print(Style.DIM + Fore.CYAN + "welcome to Uncensored library\n\t1) See all books\n\t2) Lend or return book\n\t3) Request a book\n\t4) Find book online\n\t5) Account info\n\t0) Log Out")
                self.a = int(input())
                if self.a not in [1, 2, 3, 4, 5, 0]:
                    print(Style.DIM + Fore.RED + "Please choose valid option from above...")
                else:
                    if self.a == 1:
                        self.all_books()
                    elif self.a == 2:
                        self.lend_return()
                    elif self.a == 3:
                        self.request()
                    elif self.a == 4:
                        self.find_book_online()
                    elif self.a == 5:
                        self.acc_info()
                    elif self.a == 0:
                        Log.end()
                    ws.save("Database.xlsx")
                    ws = openpyxl.load_workbook("Database.xlsx")
        except Exception as e:
            print(Style.DIM + Fore.RED + f"Error: {e}")
            self.welcome()
    # all_books, lend_return, request, find_book_online function are available to user.
    # another function which will show current status of user's book name, lending due date and lending date and charges.
    # call logout function from Log class


class Log:

    def __init__(self, id_, pass_):
        self.id_ = id_
        self.pass_ = pass_
        self.welcome()

    def login(self):
        global log_status, user_id
        while True:
            try:
                print(Style.DIM + Fore.GREEN + "Enter your id: ", end="")
                self.id_ = int(input())
                print(Style.DIM + Fore.GREEN + "Enter your password: ", end="")
                self.pass_ = input()
                for i in range(1, ws_u.max_row+1):
                    if str(ws_u.cell(i, 1).value) == str(self.id_) and str(ws_u.cell(i, 11).value) == self.pass_:
                        log_status = "Logged in"
                        user_id = self.id_
                        print(Style.DIM + Fore.CYAN + "Log in Successful...\n")
                        return log_status
                    else:
                        continue
                print(Style.DIM + Fore.RED + "Id or Password Incorrect...")
                break
            except Exception as e:
                print(Style.DIM + Fore.RED + f"Enter proper values...\nERROR: {e}")

    @staticmethod
    def register():
        while True:
            try:
                r = Reg()
                r.first_name()
            except Exception as e:
                print(Style.DIM + Fore.RED + f"Enter proper values...\nERROR: {e}")

    @staticmethod
    def end():
        global log_status
        print("Bye.....")
        print("Logging out from the server", end="")
        i = 0.5
        e = random.randint(1, 3)
        while i != e:
            print(".", end="")
            time.sleep(i)
            i += 0.5
        log_status = "Logged out"
        print("\nLogged out..")
        try:
            ws.save("Database.xlsx")
            ws.close()
            exit()

        except Exception as e:
            print("Error!!, ", e)
            exit()

    def welcome(self):
        while log_status != "Logged in":
            try:
                print(Style.DIM + Fore.CYAN + "welcome to Uncensored library\n\t1) Login\n\t2) Register\n\t3) Exit")
                a = int(input())
                if a not in [1, 2, 3]:
                    print(Style.DIM + Fore.RED + "Please choose valid option from above...")
                else:
                    if a == 1:
                        self.login()
                    elif a == 2:
                        self.register()
                    else:
                        self.end()
            except Exception as e:
                print(Style.DIM + Fore.RED + f"Value should be number...\nERROR: {e}")


class Reg:

    def first_name(self):
        global f_name
        try:
            print(Style.BRIGHT + Fore.WHITE + "Please enter your first name: ", end="")
            f_name = input()
            if f_name.isalpha():
                return self.midd_name()
            else:
                if f_name == "":
                    print(Style.DIM + Fore.RED + "First Name can not be empty")
                else:
                    print(Style.DIM + Fore.RED + "Name only contains letters...")
        except Exception as e:
            print("Error!!!, ", e)
            self.first_name()

    def midd_name(self):
        global m_name
        try:
            print(Style.BRIGHT + Fore.WHITE + "Please enter your middle name: ", end="")
            m_name = input()
            if m_name.isalpha():
                return self.last_name()
            elif m_name == "":
                m_name = "None"
                return self.last_name()
            else:
                print(Style.DIM + Fore.RED + "Name only contains letters...")
        except Exception as e:
            print("Error!!!, ", e)
            self.midd_name()

    def last_name(self):
        global l_name
        try:
            print(Style.BRIGHT + Fore.WHITE + "Please enter your last name: ", end="")
            l_name = input()
            if l_name.isalpha():
                return self.d_o_b()
            else:
                if l_name == "":
                    print(Style.DIM + Fore.RED + "Last Name can not be empty")
                else:
                    print(Style.DIM + Fore.RED + "Name only contains letters...")
        except Exception as e:
            print("Error!!!, ", e)
            self.last_name()

    def d_o_b(self):
        global dob
        dob = ""
        try:
            print(Style.BRIGHT + Fore.WHITE + "Enter date of birth in (YYYY-MM-DD) format: ", end="")
            self.dob = input()
            dt = datetime.date.today()
            if (int(self.dob[0:4]) <= int(dt.year)) and (len(self.dob) == 10):
                if (self.dob[4] == "-") and (self.dob[7] == "-"):
                    if int(self.dob[5:7]) <= 12:
                        if (int(self.dob[8:10]) <= 30) or (int(self.dob[8:10]) <= 31):
                            dob = self.dob
                            return self.age_calc()
                        else:
                            print(Style.DIM + Fore.RED + "Please enter valid Birth Date...")
                    else:
                        print(Style.DIM + Fore.RED + "Please enter valid Birth Month...")
                else:
                    print(Style.DIM + Fore.RED + "Enter date of birth with valid format")
            else:
                if len(dob) != 10:
                    print(Style.DIM + Fore.RED + "Enter date of birth with valid format")
                else:
                    print(Style.DIM + Fore.RED + "Please enter valid Birth Year...")
        except Exception as e:
            print("Error!!!, ", e)
            self.d_o_b()

    def age_calc(self):
        global age, dob
        try:
            dt = datetime.date.today()
            age = int(dt.year) - int(dob[0:4]) - (((int(dt.month)), int(dt.day)) < ((int(dob[5:7])), int(dob[8:10])))  # Subtract today's year to birth year, then compare current month & birthdate to birth month and birthdate
            return self.gend()
        except Exception as e:
            print("Error!!!, ", e)
            self.age_calc()

    def gend(self):
        global gender
        try:
            print(Style.BRIGHT + Fore.WHITE + "Enter your gender (M)-Male, (F)-Female, (U)-Unspecified: ", end="")
            gender = input()
            if gender.upper() in ("M", "F", "U"):
                return self.mail()
            else:
                print(Style.DIM + Fore.RED + "Please specify proper gender...")
        except Exception as e:
            print("Error!!!, ", e)
            self.gend()

    def mail(self):
        global email
        try:
            email_suffix = ["gmail.com", "yahoo.com", "hotmail.com", "aol.com", "hotmail.co.uk", "hotmail.fr", "msn.com",
                            "yahoo.fr", "wanadoo.fr", "orange.fr", "comcast.net", "yahoo.co.uk", "yahoo.com.br",
                            "yahoo.co.in", "live.com", "rediffmail.com",
                            "free.fr", "gmx.de", "web.de", "yandex.ru", "ymail.com", "libero.it", "outlook.com",
                            "outlook.ca", "uol.com.br", "bol.com.br", "mail.ru", "cox.net", "hotmail.it", "sbcglobal.net",
                            "sfr.fr", "live.fr", "verizon.net", "live.co.uk", "googlemail.com",
                            "yahoo.es", "ig.com.br", "live.nl", "bigpond.com", "terra.com.br", "yahoo.it", "neuf.fr",
                            "yahoo.de", "alice.it", "rocketmail.com", "att.net", "laposte.net", "facebook.com",
                            "bellsouth.net", "yahoo.in", "hotmail.es", "charter.net",
                            "yahoo.ca", "yahoo.com.au", "rambler.ru", "hotmail.de", "tiscali.it", "shaw.ca", "yahoo.co.jp",
                            "sky.com", "earthlink.net", "optonline.net", "freenet.de", "t-online.de", "aliceadsl.fr",
                            "virgilio.it", "home.nl", "qq.com", "telenet.be",
                            "me.com", "yahoo.com.ar", "tiscali.co.uk", "yahoo.com.mx", "voila.fr", "gmx.net", "mail.com",
                            "planet.nl", "tin.it", "live.it", "ntlworld.com", "arcor.de", "yahoo.co.id", "frontiernet.net",
                            "hetnet.nl", "live.com.au", "yahoo.com.sg",
                            "zonnet.nl", "club-internet.fr", "juno.com", "optusnet.com.au", "blueyonder.co.uk",
                            "bluewin.ch", "skynet.be", "sympatico.ca", "windstream.net", "mac.com", "centurytel.net",
                            "chello.nl", "live.ca", "aim.com", "bigpond.net.au"]
            print(Style.BRIGHT + Fore.WHITE + "Enter your email: ", end="")
            email = input()
            if "@" in email:
                if email.split("@", 2)[1] in email_suffix:
                    sp_c = ["!", "@", "#", "$", "%", "^", "&", "*", "(", ")", "_", "+"]
                    while True:
                        if not (any(em in email.split("@", 2)[0] for em in sp_c)):
                            return self.phone()
                        else:
                            print(Style.DIM + Fore.RED + "Email Suffix can't contain special characters...")
                else:
                    print(Style.DIM + Fore.RED + "Please specify email with proper prefix and suffix...")
            else:
                if "@" not in email:
                    print(Style.DIM + Fore.RED + "Email contains @ as suffix...")
        except Exception as e:
            print("Error!!!, ", e)
            self.mail()

    def phone(self):
        global pn
        try:
            print(Style.BRIGHT + Fore.WHITE + "Enter country code with (+): ", end="")
            p_cd = input()
            cde = {}
            i = 1
            while i < (wka.max_row - 1):
                cde[wka.cell(i, 1).value] = wka.cell(i, 3).value
                i += 1
            if p_cd in cde.values():
                while True:
                    print(Style.BRIGHT + Fore.WHITE + "Enter your 10 digit phone number: ", end="")
                    p_num = input()
                    if len(p_num) == 10:
                        if p_num.isnumeric():
                            pn = str(p_cd) + str(p_num)
                            return self.city()
                    else:
                        print(Style.DIM + Fore.RED + "Please enter valid phone number...")
            else:
                print(Style.DIM + Fore.RED + "Please enter valid country code...")
        except Exception as e:
            print("Error!!!, ", e)
            self.phone()

    def city(self):
        global u_city
        try:
            print(Style.BRIGHT + Fore.WHITE + "Enter your city name: ", end="")
            u_city = input()
            if u_city.isalpha():
                return self.addr()
            else:
                print(Style.DIM + Fore.RED + "Please enter valid city name...")
        except Exception as e:
            print("Error!!!, ", e)
            self.city()

    def addr(self):
        global address
        try:
            print(Style.BRIGHT + Fore.WHITE + "Enter your address: ", end="")
            address = input()
            if address != " " or address != "":
                return self.pin()
            else:
                print(Style.DIM + Fore.RED + "Please enter valid address...")
        except Exception as e:
            print("Error!!!, ", e)
            self.addr()

    def pin(self):
        global acc_pin
        try:
            print(Style.BRIGHT + Fore.WHITE + "Enter 4 digit access pin for security purpose: ", end="")
            acc_pin = input()
            if acc_pin.isnumeric():
                if len(acc_pin) == 4:
                    if acc_pin[0] != acc_pin[1] != acc_pin[2] != acc_pin[3]:
                        return self.GenId()
                    else:
                        print(Style.DIM + Fore.RED + "The pin should be odd not in the sequential order like 0000,1111...")
                else:
                    print(Style.DIM + Fore.RED + "The pin should be 4 digit only...")
            else:
                print(Style.DIM + Fore.RED + "The pin should contains only digit...")
        except Exception as e:
            print("Error!!!, ", e)
            self.pin()

    def GenId(self):
        global AcId
        try:
            while True:
                AcId = random.randint(1111111111, 9999999999)
                i = 2
                acc_list = []
                while i < wsa.max_row:
                    acc_no_database = wsa.cell(i, 1)
                    acc_list.append(acc_no_database)
                    i += 1
                if AcId not in acc_list:
                    return self.details(f_name, m_name, l_name, dob, age, gender, email, pn, acc_pin, AcId, u_city, address)
        except Exception as e:
            print("Error!!!, ", e)

    def details(self, f_name, m_name, l_name, dob, age, gender, email, pn, acc_pin, AcId, u_city, address):
        try:
            print(Style.BRIGHT + Fore.BLUE + "Please verify all of your details:")
            print(Style.BRIGHT + Fore.BLUE + "First Name: " + Style.DIM + Fore.YELLOW + f_name)
            print(Style.BRIGHT + Fore.BLUE + "Middle Name: " + Style.DIM + Fore.YELLOW + m_name)
            print(Style.BRIGHT + Fore.BLUE + "Last Name: " + Style.DIM + Fore.YELLOW + l_name)
            print(Style.BRIGHT + Fore.BLUE + "Date of Birth: " + Style.DIM + Fore.YELLOW + dob)
            print(Style.BRIGHT + Fore.BLUE + "Age: " + Style.DIM + Fore.YELLOW + str(age))
            print(Style.BRIGHT + Fore.BLUE + "Gender: " + Style.DIM + Fore.YELLOW + gender)
            print(Style.BRIGHT + Fore.BLUE + "City: " + Style.DIM + Fore.YELLOW + u_city)
            print(Style.BRIGHT + Fore.BLUE + "Gender: " + Style.DIM + Fore.YELLOW + address)
            print(Style.BRIGHT + Fore.BLUE + "Email: " + Style.DIM + Fore.YELLOW + email)
            print(Style.BRIGHT + Fore.BLUE + "Phone number: " + Style.DIM + Fore.YELLOW + pn)
            print(Style.BRIGHT + Fore.BLUE + "User access pin: " + Style.DIM + Fore.YELLOW + acc_pin)
            print(Style.BRIGHT + Fore.BLUE + "Account ID: " + Style.DIM + Fore.YELLOW + str(AcId))
            print(Style.DIM + Fore.RED + "Please ensure that all the details that you have provided must verify with your identity\n")
            return self.verify()
        except Exception as e:
            print("Error!!!, ", e)
            self.details(f_name, m_name, l_name, dob, age, gender, email, pn, acc_pin, AcId, u_city, address)

    # Ask if the details are correct, y== go to next step, no=> ask to choose which step is wrong
    def verify(self):
        try:
            print(Style.BRIGHT + Fore.GREEN + "Please choose (Y) or (N) if the above information are  correct or not: ", end="")
            Y_N = input()
            if Y_N.upper() == "Y":
                self.t_c()
                return self.t_c()  # move to Terms and conditions
            elif Y_N.upper() == "N":
                return Log(0, 0)
            else:
                print(Style.BRIGHT + Fore.RED + "Please enter valid input...")
        except Exception as e:
            print("Error!!!, ", e)
            self.verify()

    # Print statement which ask user to accept the Terms and Conditions for opening an account with swiss bank. If y== store the details in the Excel and print message creating account... and after that display message which shows that we are reviewing your account, and you will get updated in 2-3 business days. If no== Print message, sorry you are ineligible to open an account with us as you are not agreeing with our T&C
    def t_c(self):
        try:
            global T_C
            print(Style.DIM + Fore.RED + "\nPlease press (Y) to agree with our " + Style.BRIGHT + Fore.GREEN + "TERMS AND CONDITION" + Style.DIM + Fore.RED + " OR enter (N) to deny " + Style.RESET_ALL)
            print(Style.DIM + Fore.RED + "NOTE: By pressing (N) your account will not be created and all of the information you have entered above will be erased " + Style.RESET_ALL, end="")
            T_C = input()
            if T_C.upper() == "Y":
                self.saving()  # proceed further
            elif T_C.upper() == "N":
                print(Style.BRIGHT + Fore.GREEN + "Are you sure that you are not agreeing with our T&C's: " + Style.RESET_ALL, end="")
                D_V = input()
                if D_V.upper() == "Y":
                    print("Sorry! As you are not agreeing with our terms and conditions we can not proceed further.")  # ask which details of are false by showing the menu
                    return T_C  # cancel account creating process
                elif D_V.upper() == "N":
                    print("Account creation under progress")
                    T_C = "Y"
                    self.saving()  # continue the account creation process
                else:
                    print(Style.DIM + Fore.RED + "Please enter valid input...")
            else:
                print(Style.DIM + Fore.RED + "Please enter valid input...")
        except Exception as e:
            print("Error!!!, ", e)
            self.t_c()

    @staticmethod
    def saving():
        global f_name, m_name, l_name, dob, age, gender, email, pn, acc_pin, AcId, u_city, address
        try:
            mx_row = wsa.max_row
            cur_row = mx_row + 1
            ws_u.cell(cur_row, 1).value = AcId  # Account no
            ws_u.cell(cur_row, 2).value = f_name
            ws_u.cell(cur_row, 3).value = m_name
            ws_u.cell(cur_row, 4).value = l_name
            ws_u.cell(cur_row, 5).value = dob
            ws_u.cell(cur_row, 6).value = age
            ws_u.cell(cur_row, 7).value = gender
            ws_u.cell(cur_row, 8).value = address
            ws_u.cell(cur_row, 9).value = u_city
            ws_u.cell(cur_row, 10).value = pn
            ws_u.cell(cur_row, 11).value = acc_pin
            ws_u.cell(cur_row, 12).value = email
            ws_u.cell(cur_row, 17).value = f"Your Account has been created on {datetime.date.today()}."  # Account Activity
            ws_u.cell(cur_row, 18).value = f"{datetime.date.today()}"  # Add account opening date and time.
            ws_u.cell(cur_row, 19).value = "0"
            ws.save("Database.xlsx")
            print(Style.BRIGHT + Fore.CYAN + "Hurray!, Your account have been created.")
            return Log(0, 0)
        except Exception as e:
            print("Error!!, ", e)


if __name__ == '__main__':
    a = threading.Thread(target=Log, args=(0, 0), daemon=True)

    def check():
        while not b.daemon:
            global log_status, user_id, Authorize
            if log_status == "Logged in":
                if user_id in Authorize:
                    return LibM(user_id, log_status)
                elif user_id not in Authorize:
                    return LibU(user_id, log_status)

    b = threading.Thread(target=check, daemon=False)

    a.start()
    b.start()
